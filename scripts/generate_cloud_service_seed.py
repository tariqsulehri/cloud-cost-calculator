#!/usr/bin/env python3
"""Generate the backend cloud service seed from Cloud_Services_Mapping.xlsx."""

from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Cloud_Services_Mapping.xlsx"
OUTPUT = ROOT / "backend/src/database/cloudServices.seed.ts"

PROVIDERS = ("azure", "aws", "gcp")

LEGACY_KEYS = {
    "Virtual Machines": "compute",
    "Kubernetes Service (AKS)": "kubernetes",
    "Database for PostgreSQL": "database.postgresql",
    "Redis Cache": "cache.redis",
    "Managed Disks": "block_storage",
    "Azure CDN": "cdn",
    "Application Gateway": "load_balancer.http_s",
    "Load Balancer": "load_balancer.tcp",
    "Service Bus": "queue",
    "Azure Monitor": "monitoring",
    "Virtual Network": "network.private",
    "Azure Backup": "backup",
    "Azure Functions": "serverless",
    "Defender for Cloud": "security",
}

PROVIDER_NAMESPACE_HINTS = {
    "AI + Machine Learning": "Microsoft.CognitiveServices",
    "Analytics": "Microsoft.Analytics",
    "Compute": "Microsoft.Compute",
    "Databases": "Microsoft.DBforPostgreSQL",
    "Development": "Microsoft.DevCenter",
    "Identity + Security": "Microsoft.Security",
    "Integration": "Microsoft.ServiceBus",
    "IoT + MR": "Microsoft.Devices",
    "Management + Governance": "Microsoft.Insights",
    "Media + Comms": "Microsoft.Media",
    "Migration": "Microsoft.Migrate",
    "Networking": "Microsoft.Network",
    "Storage": "Microsoft.Storage",
}

AZURE_NAMESPACE_HINTS = {
    "App Service": "Microsoft.Web",
    "Azure Functions": "Microsoft.Web",
    "Container Apps": "Microsoft.App",
    "Container Instances": "Microsoft.ContainerInstance",
    "Container Registry": "Microsoft.ContainerRegistry",
    "Kubernetes Service (AKS)": "Microsoft.ContainerService",
    "Redis Cache": "Microsoft.Cache",
    "Managed Redis": "Microsoft.Cache",
    "Cosmos DB": "Microsoft.DocumentDB",
    "Database for MySQL": "Microsoft.DBforMySQL",
    "Database for PostgreSQL": "Microsoft.DBforPostgreSQL",
    "SQL Database": "Microsoft.Sql",
    "SQL Managed Instance": "Microsoft.Sql",
    "Key Vault": "Microsoft.KeyVault",
    "Logic Apps": "Microsoft.Logic",
    "Event Grid": "Microsoft.EventGrid",
    "Event Hubs": "Microsoft.EventHub",
    "API Management": "Microsoft.ApiManagement",
    "Azure Monitor": "Microsoft.Insights",
    "Log Analytics": "Microsoft.OperationalInsights",
    "Azure Backup": "Microsoft.RecoveryServices",
    "Site Recovery": "Microsoft.RecoveryServices",
    "Azure CDN": "Microsoft.Cdn",
    "Front Door": "Microsoft.Cdn",
}


def ascii_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\u2026", "...")
    text = text.replace("\u2013", "-").replace("\u2014", "-")
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", text).strip()


def slug(value: str) -> str:
    value = ascii_text(value).lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return re.sub(r"-+", "-", value).strip("-")


def mapping_status(name: str) -> str:
    normalized = name.lower()
    if not normalized or normalized.startswith("no direct"):
        return "no_direct_equivalent"
    if "no direct" in normalized:
        return "manual_review"
    return "mapped"


def component_type(category: str, azure_name: str) -> str:
    text = f"{category} {azure_name}".lower()

    if "redis" in text or "cache" in text:
        return "cache"
    if "kubernetes" in text or "aks" in text:
        return "kubernetes"
    if "function" in text or "serverless" in text or "logic apps" in text or "container apps" in text:
        return "serverless"
    if "application gateway" in text or "load balancer" in text or "traffic manager" in text:
        return "load_balancer"
    if "cdn" in text or "front door" in text:
        return "cdn"
    if "service bus" in text or "queue" in text or "event grid" in text or "event hubs" in text or "notification hubs" in text:
        return "queue"
    if "backup" in text or "site recovery" in text:
        return "backup"
    if "monitor" in text or "log analytics" in text or "application insights" in text:
        return "monitoring"
    if "defender" in text or "key vault" in text or "security" in text or "sentinel" in text or "identity" in text or "entra" in text:
        return "security"
    if "managed disk" in text or "elastic san" in text:
        return "block_storage"
    if "file" in text and category == "Storage":
        return "file_storage"
    if "blob" in text or "data lake" in text:
        return "object_storage"
    if category == "Storage":
        return "storage"
    if category == "Databases":
        return "database"
    if category == "Networking":
        return "network"
    if category == "Compute":
        return "compute"
    return "unknown"


def provider_namespace(category: str, azure_name: str) -> str | None:
    return AZURE_NAMESPACE_HINTS.get(azure_name) or PROVIDER_NAMESPACE_HINTS.get(category)


def pricing_name(provider: str, canonical_name: str) -> str | None:
    if mapping_status(canonical_name) == "no_direct_equivalent":
        return None
    if provider == "azure":
        return canonical_name.removeprefix("Azure ").removeprefix("Microsoft ")
    return canonical_name.split(" / ")[0]


def aliases(category: str, azure: str, aws: str, gcp: str) -> list[str]:
    values = {category, azure, aws, gcp}
    for name in (azure, aws, gcp):
        if mapping_status(name) == "no_direct_equivalent":
            continue
        simplified = re.sub(r"\([^)]*\)", "", name)
        simplified = simplified.replace("Azure ", "").replace("Amazon ", "").replace("Google ", "")
        for part in re.split(r"/|,", simplified):
            part = ascii_text(part)
            if len(part) >= 3:
                values.add(part)
    return sorted(value for value in values if value and mapping_status(value) != "no_direct_equivalent")


def service_key(category: str, azure_name: str) -> str:
    return LEGACY_KEYS.get(azure_name) or f"{slug(category)}.{slug(azure_name)}"


def row_to_seed(category: str, azure: str, aws: str, gcp: str) -> dict[str, object]:
    component = component_type(category, azure)
    providers = {
        "azure": {
            "canonicalName": azure,
            "providerNamespace": provider_namespace(category, azure),
            "pricingServiceName": pricing_name("azure", azure),
            "serviceFamily": category,
            "mappingStatus": mapping_status(azure),
        },
        "aws": {
            "canonicalName": aws or "No direct equivalent",
            "providerNamespace": None,
            "pricingServiceName": pricing_name("aws", aws),
            "serviceFamily": category,
            "mappingStatus": mapping_status(aws),
        },
        "gcp": {
            "canonicalName": gcp or "No direct equivalent",
            "providerNamespace": None,
            "pricingServiceName": pricing_name("gcp", gcp),
            "serviceFamily": category,
            "mappingStatus": mapping_status(gcp),
        },
    }

    for provider in PROVIDERS:
        status = providers[provider]["mappingStatus"]
        if status == "no_direct_equivalent":
            providers[provider]["notes"] = "No direct equivalent in the source mapping."
        elif status == "manual_review":
            providers[provider]["notes"] = "Source mapping includes a no-direct note. Review before client proposal."

    return {
        "serviceKey": service_key(category, azure),
        "componentType": component,
        "sourceCategory": category,
        "mappingStatus": "mapped",
        "aliases": aliases(category, azure, aws, gcp),
        "providers": providers,
    }


def main() -> None:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    sheet = workbook["Cloud Services Mapping"]
    seeds: list[dict[str, object]] = []
    current_category = ""

    for category_raw, azure_raw, aws_raw, gcp_raw in sheet.iter_rows(min_row=2, values_only=True):
        category = ascii_text(category_raw) or current_category
        azure = ascii_text(azure_raw)
        aws = ascii_text(aws_raw)
        gcp = ascii_text(gcp_raw)
        if category and not azure and not aws and not gcp:
            current_category = category
            continue
        if not azure:
            continue
        current_category = category
        seeds.append(row_to_seed(category, azure, aws, gcp))

    duplicate_keys = sorted({seed["serviceKey"] for seed in seeds if [row["serviceKey"] for row in seeds].count(seed["serviceKey"]) > 1})
    if duplicate_keys:
        raise ValueError(f"Duplicate service keys generated: {duplicate_keys}")

    body = json.dumps(seeds, indent=2)
    OUTPUT.write_text(
        "import type { SeedService } from './CloudCatalogDatabase.js';\n\n"
        "// Generated from Cloud_Services_Mapping.xlsx by scripts/generate_cloud_service_seed.py.\n"
        f"export const cloudServiceSeeds: SeedService[] = {body};\n",
        encoding="utf-8",
    )
    print(f"Generated {len(seeds)} service mappings at {OUTPUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
